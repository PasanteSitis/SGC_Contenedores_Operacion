#!/usr/bin/env python3
# coding: utf-8
"""
build_catalog.py

Genera/actualiza catalogo_grouped.xlsx y catalog_index.json a partir de metadata.yml.
"""
from __future__ import annotations
import os, sys, json, yaml, re, unicodedata, tempfile
from datetime import datetime, timezone
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception as e:
    raise

METADATA_FILENAME = "metadata.yml"
CATALOG_FILENAME = "catalogo_grouped.xlsx"
INDEX_FILENAME = "catalog_index.json"
DEFAULT_BASE_COLS = [
    "id","nombre","area","tipo","responsable","estado",
    "fecha_creacion","ultima_revision","ruta_archivo","sidecar_dir","scan_timestamp","key","directorio_general"
]

def log(*a):
    print(datetime.now(timezone.utc).isoformat(), "-", *a, flush=True)

def normalize_header(h):
    if h is None: return ""
    s = str(h).strip().replace("_"," ").replace("-"," ")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().strip()
    s = re.sub(r'\s+', ' ', s)
    return s

def candidate_column_name(col):
    h = normalize_header(col)
    mapping = {
        "id": ["id","identificador","identificador unico","identificador único"],
        "nombre": ["nombre","documento","file","file name","nombre del documento"],
        "area": ["area","departamento","departamento/proceso","area del documento"],
        "tipo": ["tipo","categoria","categoria documental","clase"],
        "responsable": ["responsable","autor","owner","encargado"],
        "estado": ["estado","situación","situacion"],
        "fecha_creacion": ["fecha creación","fecha creacion","created","creation date","fecha"],
        "ultima_revision": ["ultima revision","última revisión","ultima revisión","last modified","last update"]
    }
    for canonical, aliases in mapping.items():
        for a in aliases:
            if a in h:
                return canonical
    return None

def load_yaml(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = yaml.safe_load(f)
            if content is None:
                return []
            if isinstance(content, dict):
                return [content]
            if isinstance(content, list):
                return content
            return []
    except Exception as e:
        log("ERROR leyendo YAML", path, ":", e)
        return []

def first_level_under_root(some_path, root):
    root = os.path.abspath(root)
    some_path = os.path.abspath(some_path)
    if not some_path.startswith(root):
        return os.path.basename(root)
    rel = os.path.relpath(some_path, root)
    parts = rel.split(os.sep)
    if not parts or parts[0] == ".":
        return os.path.basename(root)
    return parts[0]

def ancestor_tokens(path, root):
    p = os.path.dirname(os.path.abspath(path)) if os.path.isfile(path) else os.path.abspath(path)
    tokens = []
    root_abs = os.path.abspath(root)
    while True:
        if p == root_abs or len(p) < len(root_abs):
            break
        name = os.path.basename(p)
        if name:
            tokens.append(name)
        parent = os.path.dirname(p)
        if parent == p:
            break
        p = parent
    return tokens

def build_rows_from_sidecars(root):
    rows = []
    root = os.path.abspath(root)
    for dirpath, dirnames, filenames in os.walk(root):
        IGNORE_DIRS = {'.git', '.github', '__pycache__', '.venv', 'venv', 'node_modules', '.idea', '.vscode'}
        dirnames[:] = [d for d in dirnames if d not in IGNORE_DIRS and not d.startswith(".")]
        if METADATA_FILENAME not in filenames:
            continue
        meta_path = os.path.join(dirpath, METADATA_FILENAME)
        items = load_yaml(meta_path)
        for item in items:
            if not isinstance(item, dict):
                continue
            id_ = item.get("id","") or ""
            nombre = item.get("nombre","") or ""
            area = item.get("area","") or ""
            tipo = item.get("tipo","") or ""
            responsable = item.get("responsable","") or ""
            estado = item.get("estado","") or ""
            fecha_creacion = item.get("fecha_creacion","") or ""
            ultima_revision = item.get("ultima_revision","") or ""
            ruta_archivo = os.path.normpath(os.path.join(dirpath, nombre)) if nombre else ""
            directorio_general = first_level_under_root(dirpath, root)
            anc = ancestor_tokens(ruta_archivo or os.path.join(dirpath,nombre), root)
            found_13 = None
            for a in anc:
                low = a.lower()
                normalized = re.sub(r'[\s_\-]+','_', low)
                if normalized == "13_documentos_obsoletos":
                    found_13 = a
                    break
            if found_13:
                p = os.path.abspath(dirpath)
                matched_path = None
                while True:
                    if os.path.basename(p).lower() == found_13.lower():
                        matched_path = p
                        break
                    if p == os.path.abspath(root) or os.path.dirname(p) == p:
                        break
                    p = os.path.dirname(p)
                if matched_path:
                    directorio_general = first_level_under_root(matched_path, root)
                else:
                    directorio_general = found_13
                tipo = "Obsoletos"
            key = id_.strip() if id_ and str(id_).strip() else ruta_archivo
            rows.append({
                "key": key,
                "id": id_,
                "nombre": nombre,
                "area": area,
                "tipo": tipo,
                "responsable": responsable,
                "estado": estado,
                "fecha_creacion": fecha_creacion,
                "ultima_revision": ultima_revision,
                "ruta_archivo": ruta_archivo,
                "sidecar_dir": dirpath,
                "directorio_general": directorio_general,
                "scan_timestamp": datetime.now(timezone.utc).isoformat()
            })
    return rows

def read_existing_index(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {"order": [], "groups": {}}
    except Exception as e:
        log("Aviso: error leyendo index:", e)
        return {"order": [], "groups": {}}

def save_index_atomic(index, path):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def read_existing_catalog_dataframe(path):
    try:
        df = pd.read_excel(path, engine="openpyxl", dtype=str)
        if 'key' not in df.columns:
            if 'id' in df.columns:
                df['key'] = df['id'].fillna("")
                if 'ruta_archivo' in df.columns:
                    df.loc[df['key']=="", 'key'] = df.loc[df['key']=="", 'ruta_archivo']
            elif 'ruta_archivo' in df.columns:
                df['key'] = df['ruta_archivo']
            else:
                df['key'] = df.index.astype(str)
        df = df.astype(object).where(pd.notnull(df), None)
        return df
    except Exception:
        return None

def build_index_from_new_rows(index, rows, df_old=None):
    new_keys = {r['key']: r for r in rows}
    for gname, keys in list(index["groups"].items()):
        index["groups"][gname] = [k for k in keys if k in new_keys]
        if not index["groups"].get(gname):
            index["groups"].pop(gname, None)
            if gname in index["order"]:
                index["order"].remove(gname)
    keys_by_group_and_type = {}
    types_in_group = {}
    for r in rows:
        g = r["directorio_general"]
        t = r.get("tipo") or ""
        keys_by_group_and_type.setdefault(g, {}).setdefault(t, []).append(r["key"])
        types_in_group.setdefault(g, [])
        if t not in types_in_group[g]:
            types_in_group[g].append(t)
    for g in keys_by_group_and_type.keys():
        if g not in index["order"]:
            index["order"].append(g)
    for g, type_map in keys_by_group_and_type.items():
        existing = index["groups"].get(g, [])
        new_order = []
        seen = set()
        prev_keys = [k for k in existing if k in new_keys]
        prev_key_to_type = {k: new_keys[k].get("tipo","") for k in prev_keys}
        for t in types_in_group.get(g, []):
            for k in prev_keys:
                if k not in seen and prev_key_to_type.get(k,"") == t:
                    new_order.append(k); seen.add(k)
            for k in type_map.get(t, []):
                if k not in seen:
                    new_order.append(k); seen.add(k)
        for k in prev_keys:
            if k not in seen:
                new_order.append(k); seen.add(k)
        for t, ks in type_map.items():
            for k in ks:
                if k not in seen:
                    new_order.append(k); seen.add(k)
        index["groups"][g] = new_order
    return index

def merge_extra_columns(df_new, df_old):
    if df_old is None:
        return df_new
    old_cols = list(df_old.columns)
    new_cols = list(df_new.columns)
    extra_cols = [c for c in old_cols if c not in new_cols]
    if not extra_cols:
        return df_new
    df_old_indexed = df_old.set_index('key', drop=False)
    df_new_indexed = df_new.set_index('key', drop=False)
    for c in extra_cols:
        df_new_indexed[c] = None
        for key, row in df_old_indexed[c].items():
            if key in df_new_indexed.index:
                df_new_indexed.at[key, c] = row
    return df_new_indexed.reset_index(drop=True)

def write_grouped_excel_atomic(path, index, rows, column_order, df_old=None):
    row_by_key = {r["key"]: r for r in rows}
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogo"
    bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    visible_cols = column_order.copy()
    row_idx = 1
    for g in index["order"]:
        header_text = f"DIRECTORIO: {g}"
        ncols = len(visible_cols) if len(visible_cols) > 0 else 1
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
        cell = ws.cell(row=row_idx, column=1, value=header_text)
        cell.font = Font(bold=True, size=12)
        cell.alignment = align_center
        row_idx += 1
        for ci, col in enumerate(visible_cols, start=1):
            ccell = ws.cell(row=row_idx, column=ci, value=col)
            ccell.font = bold
        row_idx += 1
        group_keys = index["groups"].get(g, [])
        for k in group_keys:
            if k not in row_by_key:
                continue
            r = row_by_key[k]
            for ci, col in enumerate(visible_cols, start=1):
                val = r.get(col, None) if col in r else None
                ws.cell(row=row_idx, column=ci, value=val)
            row_idx += 1
        row_idx += 1
    for i, col in enumerate(visible_cols, start=1):
        maxlen = 0
        for cell in ws[get_column_letter(i)]:
            if cell.value:
                l = len(str(cell.value))
                if l > maxlen:
                    maxlen = l
        ws.column_dimensions[get_column_letter(i)].width = min(maxlen + 2, 60)
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)
    log("Catálogo guardado:", path)

def main():
    if len(sys.argv) < 2:
        print("Uso: python build_catalog.py <ruta_raiz> [--template plantilla.xlsx] [--sheet-per-dir]")
        sys.exit(1)
    root = os.path.abspath(sys.argv[1])
    template = None
    sheet_per_dir = False
    if "--template" in sys.argv:
        try:
            template = sys.argv[sys.argv.index("--template")+1]
        except Exception:
            template = None
    if "--sheet-per-dir" in sys.argv:
        sheet_per_dir = True
    catalog_path = os.path.join(root, CATALOG_FILENAME)
    index_path = os.path.join(root, INDEX_FILENAME)
    log("Escaneando sidecars en:", root)
    rows = build_rows_from_sidecars(root)
    log("Rows encontradas:", len(rows))
    df_old = read_existing_catalog_dataframe(catalog_path)
    index = read_existing_index(index_path)
    if template and os.path.exists(template):
        column_order = []
        try:
            df_tmp = pd.read_excel(template, engine="openpyxl", nrows=0)
            headers = list(df_tmp.columns)
            for h in headers:
                c = candidate_column_name(h)
                column_order.append(c if c else str(h))
            for c in DEFAULT_BASE_COLS:
                if c not in column_order:
                    column_order.append(c)
            log("Usando plantilla columnas:", column_order)
        except Exception as e:
            log("Error leyendo plantilla:", e)
            column_order = DEFAULT_BASE_COLS.copy()
    else:
        column_order = DEFAULT_BASE_COLS.copy()
        log("Usando columnas por defecto:", column_order)
    index = build_index_from_new_rows(index, rows, df_old=df_old)
    save_index_atomic(index, index_path)
    df_new = pd.DataFrame(rows) if rows else pd.DataFrame(columns=DEFAULT_BASE_COLS)
    df_final = merge_extra_columns(df_new, df_old)
    for c in column_order:
        if c not in df_final.columns:
            df_final[c] = None
    write_grouped_excel_atomic(catalog_path, index, rows, column_order, df_old=df_old)
    log("Listo.")

if __name__ == "__main__":
    main()